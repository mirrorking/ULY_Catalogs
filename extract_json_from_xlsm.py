# extract_xlsm_to_json_corrected.py
import pandas as pd
import json
import os
import re
from openpyxl import load_workbook

def extract_xlsm_to_json_corrected(xlsm_path, json_output_path):
    """
    从xlsm文件中提取所有工作表数据到JSON文件
    数据从第3行开始，第2行是列名，忽略第1行的图片信息
    确保CODE字段始终为字符串类型
    
    Args:
        xlsm_path: xlsm文件路径
        json_output_path: JSON输出文件路径
    """
    try:
        # 读取xlsm文件
        wb = load_workbook(xlsm_path, data_only=True)
        sheet_names = wb.sheetnames
        
        json_data = {}
        
        print(f"找到 {len(sheet_names)} 个工作表:")
        
        for sheet_name in sheet_names:
            print(f"  正在处理: {sheet_name}")
            
            # 读取整个工作表
            df = pd.read_excel(
                xlsm_path, 
                sheet_name=sheet_name, 
                engine='openpyxl',
                header=None  # 不自动设置列名
            )
            
            # 检查数据形状
            print(f"    原始数据形状: {df.shape} (行×列)")
            
            # 如果数据行数少于3，跳过
            if len(df) < 3:
                print(f"    工作表 {sheet_name} 数据不足3行，跳过")
                json_data[sheet_name] = []
                continue
            
            # 第2行（索引1）是列名，从第3行（索引2）开始是数据
            # 获取列名
            column_names = []
            for col in range(df.shape[1]):
                cell_value = df.iloc[1, col]  # 第2行
                if pd.isna(cell_value):
                    column_names.append(f"Column_{col+1}")
                else:
                    column_names.append(str(cell_value).strip())
            
            print(f"    列名: {column_names}")
            
            # 提取数据（从第3行开始）
            data_start_row = 2  # 索引2对应Excel第3行
            data_df = df.iloc[data_start_row:].copy()
            data_df.columns = column_names
            
            # 重置索引
            data_df = data_df.reset_index(drop=True)
            
            # 清理数据：去除完全空白的行
            data_df = data_df.dropna(how='all')
            
            if data_df.empty:
                print(f"    工作表 {sheet_name} 无有效数据，跳过")
                json_data[sheet_name] = []
                continue
            
            # 转换NaN值为None
            data_df = data_df.where(pd.notnull(data_df), None)
            
            # 转换为字典列表，确保CODE字段是字符串
            records = []
            for index, row in data_df.iterrows():
                record = {}
                for col in data_df.columns:
                    value = row[col]
                    
                    # 特殊处理：如果列名是CODE或code，确保值是字符串
                    col_upper = str(col).upper()
                    if col_upper == 'CODE':
                        if value is not None:
                            # 转换为字符串
                            if isinstance(value, (int, float)):
                                # 处理整数：如果是整数，转换为字符串，保留前导零
                                if isinstance(value, float) and value.is_integer():
                                    # 对于整数浮点数，转换为整数再转字符串
                                    value = str(int(value))
                                else:
                                    # 对于浮点数，转换为字符串
                                    value = str(value)
                                # 去除可能的小数点
                                if '.' in value:
                                    # 检查是否可以转换为整数
                                    try:
                                        if float(value).is_integer():
                                            value = str(int(float(value)))
                                    except:
                                        pass
                            elif isinstance(value, str):
                                value = value.strip()
                                # 如果是纯数字字符串，确保格式正确
                                if re.match(r'^\d+(\.0+)?$', value):
                                    try:
                                        num = float(value)
                                        if num.is_integer():
                                            value = str(int(num))
                                    except:
                                        pass
                    
                    # 如果值是字符串，去除多余空格
                    if isinstance(value, str):
                        value = value.strip()
                    
                    record[col] = value
                
                # 确保CODE字段存在且是字符串格式
                if 'CODE' in record:
                    code_value = record['CODE']
                    if code_value is not None:
                        # 确保是字符串
                        if not isinstance(code_value, str):
                            code_value = str(code_value)
                        # 标准化CODE格式（如果是数字，补零到6位）
                        if re.match(r'^\d+$', code_value):
                            code_value = code_value.zfill(6)
                        record['CODE'] = code_value
                    else:
                        # 如果CODE为空，生成一个
                        record['CODE'] = f"{sheet_name}_{index+1:04d}"
                elif 'code' in record:
                    # 处理小写的code字段
                    code_value = record['code']
                    if code_value is not None:
                        if not isinstance(code_value, str):
                            code_value = str(code_value)
                        if re.match(r'^\d+$', code_value):
                            code_value = code_value.zfill(6)
                        record['CODE'] = code_value
                        del record['code']  # 删除小写字段
                    else:
                        record['CODE'] = f"{sheet_name}_{index+1:04d}"
                else:
                    # 如果没有CODE字段，创建一个
                    record['CODE'] = f"{sheet_name}_{index+1:04d}"
                
                # 添加Excel行号信息
                record['_excel_row'] = index + 3  # Excel行号（从3开始）
                
                records.append(record)
            
            # 添加到JSON数据中
            json_data[sheet_name] = records
            
            print(f"    工作表 {sheet_name}: 提取了 {len(records)} 条记录")
            
            # 显示前几条记录的结构
            if records:
                first_record = records[0]
                code_value = first_record.get('CODE', '无')
                code_type = type(code_value)
                print(f"    第一条记录CODE类型: {code_type}, 值: '{code_value}'")
                
                # 显示前几条记录的CODE值
                print(f"    前5条记录CODE值:")
                for i in range(min(5, len(records))):
                    code_val = records[i].get('CODE', '无')
                    print(f"      {i+1}. '{code_val}' (类型: {type(code_val)})")
        
        # 保存为JSON文件
        with open(json_output_path, 'w', encoding='utf-8') as f:
            json.dump(json_data, f, ensure_ascii=False, indent=2, default=str)
        
        print(f"\n✅ JSON文件已保存到: {json_output_path}")
        print(f"📊 总工作表数: {len(json_data)}")
        
        # 显示统计信息
        total_records = sum(len(records) for records in json_data.values())
        print(f"📊 总记录数: {total_records}")
        
        # 检查CODE字段类型
        print("\n🔍 检查CODE字段类型:")
        for sheet_name, records in json_data.items():
            if records:
                code_types = {}
                for record in records[:10]:  # 只检查前10条记录
                    code = record.get('CODE')
                    if code is not None:
                        code_type = type(code).__name__
                        code_types[code_type] = code_types.get(code_type, 0) + 1
                
                if code_types:
                    type_str = ", ".join([f"{k}: {v}" for k, v in code_types.items()])
                    print(f"  {sheet_name}: {type_str}")
        
        return json_data
        
    except Exception as e:
        print(f"❌ 处理文件时出错: {e}")
        import traceback
        traceback.print_exc()
        return None

def inspect_worksheet_structure_detail(xlsm_path, sheet_name=None, num_rows=10):
    """
    详细检查工作表结构
    
    Args:
        xlsm_path: xlsm文件路径
        sheet_name: 指定工作表名称（可选）
        num_rows: 显示的行数
    """
    try:
        wb = load_workbook(xlsm_path, data_only=True)
        
        if sheet_name:
            if sheet_name in wb.sheetnames:
                sheet_names = [sheet_name]
            else:
                print(f"❌ 工作表 '{sheet_name}' 不存在")
                return
        else:
            sheet_names = wb.sheetnames
        
        for name in sheet_names:
            print(f"\n📄 工作表: {name}")
            
            # 读取原始数据（不设置列名）
            df = pd.read_excel(xlsm_path, sheet_name=name, engine='openpyxl', header=None, nrows=num_rows+2)
            
            print(f"  读取的行数: {len(df)}")
            print(f"  列数: {df.shape[1]}")
            
            # 显示前几行数据
            print("  前5行原始数据:")
            for i in range(min(5, len(df))):
                row_data = []
                for j in range(min(10, df.shape[1])):  # 只显示前10列
                    cell_value = df.iloc[i, j]
                    if pd.isna(cell_value):
                        row_data.append("(空)")
                    else:
                        display_value = str(cell_value)[:30]
                        # 显示值和类型
                        row_data.append(f"{display_value} ({type(cell_value).__name__})")
                print(f"    第{i+1}行: {row_data}")
            
            print("-" * 60)
            
    except Exception as e:
        print(f"❌ 检查工作表结构时出错: {e}")

def extract_with_custom_skip_rows(xlsm_path, json_output_path, skip_rows=2):
    """
    使用跳过行数的自定义提取方法，确保CODE字段是字符串
    
    Args:
        xlsm_path: xlsm文件路径
        json_output_path: JSON输出文件路径
        skip_rows: 跳过的行数（默认2行）
    """
    try:
        wb = load_workbook(xlsm_path, data_only=True)
        sheet_names = wb.sheetnames
        
        json_data = {}
        
        print(f"找到 {len(sheet_names)} 个工作表:")
        print(f"跳过前 {skip_rows} 行（包含列名行）")
        
        for sheet_name in sheet_names:
            print(f"  正在处理: {sheet_name}")
            
            try:
                # 读取工作表，跳过前skip_rows行
                df = pd.read_excel(
                    xlsm_path, 
                    sheet_name=sheet_name, 
                    engine='openpyxl',
                    skiprows=skip_rows  # 跳过行数
                )
                
                if df.empty:
                    print(f"    工作表 {sheet_name} 为空，跳过")
                    json_data[sheet_name] = []
                    continue
                
                print(f"    读取到 {len(df)} 行数据")
                print(f"    列名: {list(df.columns)}")
                
                # 转换NaN值为None
                df = df.where(pd.notnull(df), None)
                
                # 转换为字典列表，确保CODE字段是字符串
                records = []
                for index, row in df.iterrows():
                    record = {}
                    for col in df.columns:
                        value = row[col]
                        
                        # 特殊处理CODE字段
                        col_str = str(col).upper()
                        if col_str == 'CODE':
                            if value is not None:
                                # 确保CODE是字符串
                                if isinstance(value, (int, float)):
                                    if isinstance(value, float) and value.is_integer():
                                        value = str(int(value))
                                    else:
                                        value = str(value)
                                    # 标准化数字CODE
                                    if re.match(r'^\d+(\.0+)?$', value):
                                        try:
                                            num = float(value)
                                            if num.is_integer():
                                                value = str(int(num)).zfill(6)
                                        except:
                                            pass
                                elif isinstance(value, str):
                                    value = value.strip()
                                    if re.match(r'^\d+$', value):
                                        value = value.zfill(6)
                        
                        # 如果值是字符串，去除多余空格
                        if isinstance(value, str):
                            value = value.strip()
                        
                        record[col] = value
                    
                    # 确保CODE字段存在且是字符串
                    if 'CODE' in record:
                        code_val = record['CODE']
                        if code_val is not None:
                            if not isinstance(code_val, str):
                                code_val = str(code_val)
                            if re.match(r'^\d+$', code_val):
                                code_val = code_val.zfill(6)
                            record['CODE'] = code_val
                        else:
                            record['CODE'] = f"{sheet_name}_{index+1:04d}"
                    elif 'code' in record:
                        code_val = record['code']
                        if code_val is not None:
                            if not isinstance(code_val, str):
                                code_val = str(code_val)
                            if re.match(r'^\d+$', code_val):
                                code_val = code_val.zfill(6)
                            record['CODE'] = code_val
                            del record['code']
                        else:
                            record['CODE'] = f"{sheet_name}_{index+1:04d}"
                    else:
                        record['CODE'] = f"{sheet_name}_{index+1:04d}"
                    
                    # 添加Excel行号信息
                    record['_excel_row'] = index + skip_rows + 1
                    
                    records.append(record)
                
                json_data[sheet_name] = records
                print(f"    工作表 {sheet_name}: 提取了 {len(records)} 条记录")
                
                # 显示CODE字段类型
                if records:
                    first_code = records[0].get('CODE', '无')
                    print(f"    第一条记录CODE: '{first_code}' (类型: {type(first_code).__name__})")
                
            except Exception as e:
                print(f"    处理工作表 {sheet_name} 时出错: {e}")
                import traceback
                traceback.print_exc()
                json_data[sheet_name] = []
        
        # 保存为JSON文件
        with open(json_output_path, 'w', encoding='utf-8') as f:
            json.dump(json_data, f, ensure_ascii=False, indent=2, default=str)
        
        print(f"\n✅ JSON文件已保存到: {json_output_path}")
        
        return json_data
        
    except Exception as e:
        print(f"❌ 处理文件时出错: {e}")
        import traceback
        traceback.print_exc()
        return None

def check_code_field_types(json_path):
    """
    检查JSON文件中CODE字段的类型
    
    Args:
        json_path: JSON文件路径
    """
    try:
        with open(json_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        print("\n🔍 检查JSON文件中CODE字段类型:")
        for sheet_name, records in data.items():
            if records:
                code_types = {}
                for record in records[:5]:  # 只检查前5条记录
                    code = record.get('CODE')
                    if code is not None:
                        code_type = type(code).__name__
                        code_types[code_type] = code_types.get(code_type, 0) + 1
                
                if code_types:
                    type_info = []
                    for type_name, count in code_types.items():
                        # 获取示例值
                        example = None
                        for record in records:
                            if record.get('CODE') is not None and type(record.get('CODE')).__name__ == type_name:
                                example = record.get('CODE')
                                break
                        type_info.append(f"{type_name}: {count} (示例: '{example}')")
                    
                    print(f"  {sheet_name}: {', '.join(type_info)}")
        
        return True
    except Exception as e:
        print(f"❌ 检查CODE字段类型时出错: {e}")
        return False

def main():
    """主函数"""
    print("=" * 60)
    print("XLSM to JSON 转换工具 (修正版)")
    print("确保所有CODE字段为字符串类型")
    print("=" * 60)
    
    # 配置文件路径
    xlsm_file = input("请输入xlsm文件路径 (默认: ./商品目录/2025 ULY Catalog(添加新编码 12.24).xlsm): ").strip()
    if not xlsm_file:
        xlsm_file = "./商品目录/2025 ULY Catalog(添加新编码 12.24).xlsm"
    
    json_output = "products_data.json"
    
    # 检查文件是否存在
    if not os.path.exists(xlsm_file):
        print(f"❌ 文件 {xlsm_file} 不存在")
        print("请确保文件存在，或者输入正确的路径")
        return
    
    print(f"\n📁 输入文件: {xlsm_file}")
    print(f"📄 输出文件: {json_output}")
    
    # 首先检查文件结构
    print("\n🔍 检查文件结构...")
    inspect_worksheet_structure_detail(xlsm_file)
    
    # 选择提取方法
    print("\n🎯 选择提取方法:")
    print("1. 自动检测结构 (从第3行开始提取，推荐)")
    print("2. 指定跳过行数 (默认跳过2行)")
    print("3. 手动指定数据起始行和列名行")
    
    choice = input("请输入选择 (1/2/3): ").strip()
    
    if choice == "1":
        print("\n🚀 使用自动检测结构方法...")
        result = extract_xlsm_to_json_corrected(xlsm_file, json_output)
    elif choice == "2":
        try:
            skip_rows = int(input("请输入要跳过的行数 (默认2): ").strip() or "2")
        except:
            skip_rows = 2
        print(f"\n🚀 使用跳过 {skip_rows} 行的方法...")
        result = extract_with_custom_skip_rows(xlsm_file, json_output, skip_rows)
    elif choice == "3":
        print("\n🔧 手动指定参数...")
        try:
            header_row = int(input("列名所在行号 (默认2): ").strip() or "2")
            data_start_row = int(input("数据起始行号 (默认3): ").strip() or "3")
            
            print(f"\n🚀 使用自定义参数: 列名行={header_row}, 数据起始行={data_start_row}")
            
            # 读取文件
            wb = load_workbook(xlsm_file, data_only=True)
            sheet_names = wb.sheetnames
            
            json_data = {}
            
            for sheet_name in sheet_names:
                print(f"  处理: {sheet_name}")
                
                # 读取数据
                df = pd.read_excel(xlsm_file, sheet_name=sheet_name, engine='openpyxl', header=None)
                
                # 获取列名（从指定行）
                column_names = []
                if header_row - 1 < len(df):
                    for col in range(df.shape[1]):
                        cell_value = df.iloc[header_row-1, col]
                        if pd.isna(cell_value):
                            column_names.append(f"Column_{col+1}")
                        else:
                            column_names.append(str(cell_value).strip())
                
                # 提取数据（从指定行开始）
                data_df = df.iloc[data_start_row-1:].copy()
                data_df.columns = column_names
                
                # 清理和转换
                data_df = data_df.dropna(how='all')
                data_df = data_df.where(pd.notnull(data_df), None)
                
                # 转换为字典列表，确保CODE字段是字符串
                records = []
                for index, row in data_df.iterrows():
                    record = {}
                    for col in data_df.columns:
                        value = row[col]
                        
                        # 处理CODE字段
                        col_str = str(col).upper()
                        if col_str == 'CODE':
                            if value is not None:
                                if isinstance(value, (int, float)):
                                    if isinstance(value, float) and value.is_integer():
                                        value = str(int(value))
                                    else:
                                        value = str(value)
                                    # 标准化数字
                                    if re.match(r'^\d+(\.0+)?$', value):
                                        try:
                                            num = float(value)
                                            if num.is_integer():
                                                value = str(int(num)).zfill(6)
                                        except:
                                            pass
                                elif isinstance(value, str):
                                    value = value.strip()
                                    if re.match(r'^\d+$', value):
                                        value = value.zfill(6)
                        
                        # 如果值是字符串，去除多余空格
                        if isinstance(value, str):
                            value = value.strip()
                        
                        record[col] = value
                    
                    # 确保CODE字段存在且是字符串
                    if 'CODE' in record:
                        code_val = record['CODE']
                        if code_val is not None:
                            if not isinstance(code_val, str):
                                code_val = str(code_val)
                            if re.match(r'^\d+$', code_val):
                                code_val = code_val.zfill(6)
                            record['CODE'] = code_val
                        else:
                            record['CODE'] = f"{sheet_name}_{index+1:04d}"
                    elif 'code' in record:
                        code_val = record['code']
                        if code_val is not None:
                            if not isinstance(code_val, str):
                                code_val = str(code_val)
                            if re.match(r'^\d+$', code_val):
                                code_val = code_val.zfill(6)
                            record['CODE'] = code_val
                            del record['code']
                        else:
                            record['CODE'] = f"{sheet_name}_{index+1:04d}"
                    else:
                        record['CODE'] = f"{sheet_name}_{index+1:04d}"
                    
                    record['_excel_row'] = index + data_start_row
                    records.append(record)
                
                json_data[sheet_name] = records
                print(f"    提取了 {len(records)} 条记录")
                
                # 显示CODE类型
                if records:
                    first_code = records[0].get('CODE', '无')
                    print(f"    第一条记录CODE: '{first_code}' (类型: {type(first_code).__name__})")
            
            # 保存JSON文件
            with open(json_output, 'w', encoding='utf-8') as f:
                json.dump(json_data, f, ensure_ascii=False, indent=2, default=str)
            
            result = json_data
            print(f"\n✅ JSON文件已保存到: {json_output}")
            
        except Exception as e:
            print(f"❌ 处理时出错: {e}")
            import traceback
            traceback.print_exc()
            result = None
    else:
        print("❌ 无效选择")
        return
    
    if result:
        print("\n✅ 提取完成!")
        print(f"📊 总计: {len(result)} 个工作表")
        
        # 显示详细统计信息
        total_records = sum(len(records) for records in result.values())
        print(f"📊 总记录数: {total_records}")
        
        for sheet_name, records in result.items():
            if records:
                print(f"  - {sheet_name}: {len(records)} 条记录")
                # 显示CODE字段信息
                code_examples = []
                for i in range(min(3, len(records))):
                    code = records[i].get('CODE', '无')
                    code_type = type(code).__name__
                    code_examples.append(f"'{code}' ({code_type})")
                if code_examples:
                    print(f"    前3条CODE: {', '.join(code_examples)}")
        
        # 验证JSON文件
        if os.path.exists(json_output):
            file_size = os.path.getsize(json_output) / 1024  # KB
            print(f"\n📏 JSON文件大小: {file_size:.2f} KB")
            
            # 测试读取JSON
            try:
                with open(json_output, 'r', encoding='utf-8') as f:
                    test_data = json.load(f)
                print(f"✅ JSON文件验证通过，可正确读取")
                
                # 检查CODE字段类型
                check_code_field_types(json_output)
                
                # 显示一个示例
                first_sheet = next(iter(test_data.keys()))
                if test_data[first_sheet]:
                    print(f"\n📋 工作表 '{first_sheet}' 第一条记录示例:")
                    example = test_data[first_sheet][0]
                    for key, value in list(example.items())[:6]:  # 显示前6个字段
                        value_str = str(value)[:50]
                        if len(str(value)) > 50:
                            value_str += '...'
                        print(f"    {key}: {value_str} ({type(value).__name__})")
                
            except Exception as e:
                print(f"❌ JSON文件读取错误: {e}")
        
        print("\n🎯 HTML使用说明:")
        print("1. 将 products_data.json 放在HTML文件同级目录")
        print("2. 将图片放在 images/ 目录下，命名为 {CODE}.png/jpg/jpeg")
        print("3. CODE字段已确保为字符串类型，可直接用于图片加载")
    
    else:
        print("❌ 提取失败，请检查文件格式")

if __name__ == "__main__":
    main()